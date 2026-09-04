from pathlib import Path
import csv, random, math
from datetime import date, timedelta, datetime
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT=Path('/mnt/data')
random.seed(42)
N_RISKS=5000
N_CONTROLS=1400
N_ACCESS=12000
AS_OF=date(2026,9,1)

DEPTS=['Information Technology','Finance','Human Resources','Operations','Sales','Legal','Procurement','Customer Support']
DEPT_W=[0.25,0.14,0.10,0.16,0.10,0.07,0.08,0.10]
CATS=['Cybersecurity','Data Privacy','Regulatory Compliance','Third-Party','Operational','Financial','Business Continuity','Access Management']
ASSETS=['Web Application','Database','Cloud Service','Endpoint Fleet','ERP System','HR Platform','Payment System','Vendor Service','Network Infrastructure']
OWNERS=['A. Perera','N. Silva','K. Fernando','S. Jayasinghe','D. Senanayake','M. Wijesinghe','R. Gunawardena','T. Karunaratne']
REGS=['None','ISO 27001','NIST CSF','PCI DSS','GDPR','SOX','Local Privacy Law']
TREAT=['Not Started','Planned','In Progress','Implemented','Risk Accepted']
CONTROL_TYPES=['Preventive','Detective','Corrective']
CONTROL_DOMAINS=['Access Control','Vulnerability Management','Incident Response','Data Protection','Third-Party Risk','Business Continuity','Change Management','Security Awareness']
SYSTEMS=['ERP-PROD','CRM-CLOUD','HRIS','PAYMENTS','DATA-WAREHOUSE','CUSTOMER-PORTAL','EMAIL','DEVOPS','FILE-SHARE']
ROLES=['Standard User','Finance Analyst','HR Specialist','Developer','System Administrator','Database Administrator','Support Agent','Contractor']
COUNTRIES=['Sri Lanka','India','Singapore','United Kingdom','United States','Germany']

# helpers
def clamp(x,a,b): return max(a,min(b,x))
def weighted_choice(items, weights): return random.choices(items, weights=weights, k=1)[0]
def iso(d): return d.isoformat()

def risk_level(score):
    if score < 5.5: return 'Low'
    if score < 9.5: return 'Medium'
    if score < 14.5: return 'High'
    return 'Critical'

risk_rows=[]
for i in range(1,N_RISKS+1):
    dept=weighted_choice(DEPTS,DEPT_W); cat=random.choice(CATS); asset=random.choice(ASSETS)
    likelihood=clamp(round(random.gauss(2.7 + (cat in ['Cybersecurity','Third-Party'])*0.35,1.0)),1,5)
    impact=clamp(round(random.gauss(3.0 + (asset in ['Database','Payment System','ERP System'])*0.45,1.0)),1,5)
    criticality=clamp(round(random.gauss(3.2,1.0)),1,5)
    sensitivity=clamp(round(random.gauss(3.0 + (asset in ['Database','HR Platform','Payment System'])*0.6,1.0)),1,5)
    vulns=max(0,round(random.gammavariate(2.0,3.0)+(cat=='Cybersecurity')*4))
    crit_v=min(vulns,max(0,round(random.gammavariate(1.3,0.8))))
    incidents=max(0,round(random.gammavariate(1.1,0.75)))
    policy=max(0,round(random.gammavariate(1.2,1.1)))
    findings=max(0,round(random.gammavariate(1.4,1.25)))
    third_party=1 if cat=='Third-Party' or asset=='Vendor Service' or random.random()<0.22 else 0
    n_controls=random.randint(1,7)
    effectiveness=clamp(random.betavariate(5,2.5),0.15,0.97)
    test_pass=clamp(effectiveness+random.gauss(0,0.10),0.05,1.0)
    days_test=random.randint(3,500)
    deficiencies=max(0,round((1-effectiveness)*5+random.gauss(0,1)))
    regulatory=random.choice(REGS[1:]) if sensitivity>=4 or cat in ['Data Privacy','Regulatory Compliance','Financial'] else weighted_choice(REGS,[0.55]+[0.45/6]*6)
    days_open=random.randint(5,900)
    treatment=weighted_choice(TREAT,[0.16,0.19,0.31,0.25,0.09])
    due_days=random.randint(-180,240) if treatment not in ['Implemented','Risk Accepted'] else random.randint(-365,60)
    accepted=1 if treatment=='Risk Accepted' else 0
    inherent=likelihood*impact
    residual=inherent*(1-effectiveness)
    latent=(residual + 0.35*crit_v + 0.45*incidents + 0.28*deficiencies + 0.18*findings + 0.12*policy + 0.55*third_party + 0.22*(criticality-3) + 0.18*(sensitivity-3) + random.gauss(0,1.15))
    level=risk_level(latent)
    title=f'{cat} risk affecting {asset.lower()}'
    desc=f'Potential {cat.lower()} event may affect the {asset.lower()} used by {dept}.'
    status='Overdue' if due_days<0 and treatment not in ['Implemented','Risk Accepted'] else ('Closed' if treatment=='Implemented' else 'Open')
    row=[f'R{i:05d}',title,desc,dept,cat,asset,random.choice(OWNERS),likelihood,impact,criticality,sensitivity,vulns,crit_v,incidents,policy,findings,third_party,n_controls,round(effectiveness,3),round(test_pass,3),days_test,deficiencies,regulatory,treatment,due_days,accepted,days_open,round(inherent,2),round(residual,2),round(latent,2),level,status]
    risk_rows.append(row)

risk_headers=['risk_id','risk_title','risk_description','department','risk_category','asset_type','risk_owner','likelihood','impact','asset_criticality','data_sensitivity','vulnerability_count','critical_vulnerability_count','past_incident_count','policy_violation_count','audit_finding_count','third_party_dependency','number_of_controls','control_effectiveness','control_test_pass_rate','days_since_control_test','open_control_deficiencies','regulatory_relevance','treatment_status','treatment_due_days','risk_accepted','days_open','inherent_risk_score','residual_risk_score','adjusted_risk_score','risk_level','action_status']

control_rows=[]
for i in range(1,N_CONTROLS+1):
    domain=random.choice(CONTROL_DOMAINS); ctype=weighted_choice(CONTROL_TYPES,[0.48,0.38,0.14]); owner=random.choice(OWNERS)
    freq=weighted_choice(['Continuous','Monthly','Quarterly','Semi-Annual','Annual'],[0.18,0.20,0.30,0.16,0.16])
    automated=1 if random.random()<0.48 else 0
    design=weighted_choice(['Effective','Needs Improvement','Ineffective'],[0.72,0.20,0.08])
    operating=weighted_choice(['Effective','Needs Improvement','Ineffective','Not Tested'],[0.62,0.22,0.08,0.08])
    samples=random.choice([5,10,15,20,25,30,40])
    fails=0 if operating=='Effective' else min(samples, max(1,round(samples*random.uniform(.05,.45))))
    last_test=AS_OF-timedelta(days=random.randint(1,540))
    next_test=last_test+timedelta(days={'Continuous':30,'Monthly':30,'Quarterly':90,'Semi-Annual':180,'Annual':365}[freq])
    linked=random.choice(risk_rows)[0]
    deficiency='None' if fails==0 else weighted_choice(['Low','Medium','High','Critical'],[0.35,0.40,0.20,0.05])
    status='Overdue' if next_test<AS_OF else 'Scheduled'
    control_rows.append([f'C{i:04d}',f'{domain} Control {i:04d}',domain,ctype,owner,freq,automated,design,operating,samples,fails,iso(last_test),iso(next_test),linked,deficiency,status])
control_headers=['control_id','control_name','control_domain','control_type','control_owner','test_frequency','automated_control','design_effectiveness','operating_effectiveness','sample_size','failed_samples','last_test_date','next_test_date','linked_risk_id','deficiency_severity','test_status']

access_rows=[]
for i in range(1,N_ACCESS+1):
    uid=f'U{random.randint(1,900):04d}'; role=weighted_choice(ROLES,[.35,.10,.08,.16,.05,.03,.16,.07]); system=random.choice(SYSTEMS)
    privileged=1 if role in ['System Administrator','Database Administrator'] or (role=='Developer' and random.random()<.15) else 0
    anomaly=random.random()<0.035
    successful=max(1,round(random.gauss(12,7)))
    failed=max(0,round(random.gammavariate(1.2,1.3)))
    after=max(0,round(random.gammavariate(1.0,.7)))
    records=max(0,round(random.lognormvariate(4.5,1.0)))
    locations=weighted_choice([1,2,3],[.84,.14,.02])
    country=weighted_choice(COUNTRIES,[.72,.10,.06,.05,.04,.03])
    if anomaly:
        pattern=random.choice(['failed_logins','after_hours','data_volume','multi_location','privileged_spike'])
        if pattern=='failed_logins': failed+=random.randint(15,55)
        elif pattern=='after_hours': after+=random.randint(12,40)
        elif pattern=='data_volume': records+=random.randint(7000,30000)
        elif pattern=='multi_location': locations=random.randint(4,8)
        else: privileged=1; successful+=random.randint(60,160)
    ts=datetime.combine(AS_OF-timedelta(days=random.randint(0,90)), datetime.min.time())+timedelta(hours=random.randint(0,23),minutes=random.randint(0,59))
    access_rows.append([f'A{i:06d}',uid,random.choice(DEPTS),role,system,privileged,successful,failed,after,records,locations,country,ts.isoformat(timespec='minutes'),1 if anomaly else 0])
access_headers=['activity_id','user_id','department','user_role','system_name','privileged_access','successful_logins','failed_logins','after_hours_events','sensitive_records_accessed','distinct_locations','primary_country','activity_timestamp','synthetic_anomaly_label']

# CSV files
for name,headers,rows in [('risk_register.csv',risk_headers,risk_rows),('control_register.csv',control_headers,control_rows),('access_activity.csv',access_headers,access_rows)]:
    with open(OUT/name,'w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(headers); w.writerows(rows)

# workbook
wb=Workbook(); wb.remove(wb.active)
navy='17365D'; teal='1B7F79'; light='D9EAF7'; orange='F4B183'; red='F4CCCC'; gray='E7E6E6'; purple='D9D2E9'

def add_sheet(name,headers,rows,table_name):
    ws=wb.create_sheet(name); ws.sheet_view.showGridLines=False
    ws.append(headers)
    for row in rows: ws.append(row)
    for c in ws[1]:
        c.fill=PatternFill('solid',fgColor=navy); c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center')
    ws.freeze_panes='A2'
    tab=Table(displayName=table_name,ref=f'A1:{get_column_letter(len(headers))}{len(rows)+1}')
    tab.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True,showFirstColumn=False,showLastColumn=False)
    ws.add_table(tab)
    widths={h:min(max(len(h)+2,12),28) for h in headers}
    for idx,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(idx)].width=widths[h]
    return ws

wsr=add_sheet('Risk Register',risk_headers,risk_rows,'RiskRegisterTable')
# convert derived columns in xlsx to formulas
for r in range(2,N_RISKS+2):
    wsr.cell(r,28).value=f'=H{r}*I{r}'
    wsr.cell(r,29).value=f'=AB{r}*(1-S{r})'
    wsr.cell(r,30).value=f'=AC{r}+0.35*M{r}+0.45*N{r}+0.28*V{r}+0.18*P{r}+0.12*O{r}+0.55*Q{r}+0.22*(J{r}-3)+0.18*(K{r}-3)'
    wsr.cell(r,31).value=f'=IF(AD{r}<5.5,"Low",IF(AD{r}<9.5,"Medium",IF(AD{r}<14.5,"High","Critical")))'
for col in ['S','T']:
    for c in wsr[col][1:]: c.number_format='0.0%'
wsr.conditional_formatting.add(f'AE2:AE{N_RISKS+1}',FormulaRule(formula=['AE2="Critical"'],fill=PatternFill('solid',fgColor=red)))
wsr.conditional_formatting.add(f'AE2:AE{N_RISKS+1}',FormulaRule(formula=['AE2="High"'],fill=PatternFill('solid',fgColor=orange)))

wsc=add_sheet('Control Register',control_headers,control_rows,'ControlRegisterTable')
for col in ['L','M']:
    for c in wsc[col][1:]: c.number_format='yyyy-mm-dd'
wsa=add_sheet('Access Activity',access_headers,access_rows,'AccessActivityTable')

# README
ws=wb.create_sheet('README',0); ws.sheet_view.showGridLines=False
ws['A1']='GRC Risk Intelligence Project - Synthetic Datasets'; ws['A1'].font=Font(size=18,bold=True,color='FFFFFF'); ws['A1'].fill=PatternFill('solid',fgColor=navy); ws.merge_cells('A1:F1')
notes=[
('Purpose','Interview portfolio dataset for ML-assisted risk prioritization, control monitoring, and access anomaly detection.'),
('Data status','All people, organizations, events, risk records, and labels are synthetic. Do not treat them as real compliance evidence.'),
('As-of date',AS_OF.isoformat()),
('Risk register',f'{N_RISKS:,} records. Target: risk_level. adjusted_risk_score simulates expert judgement and is retained for transparency.'),
('Control register',f'{N_CONTROLS:,} control-testing records linked to risk IDs.'),
('Access activity',f'{N_ACCESS:,} aggregated activity records. synthetic_anomaly_label is available for evaluation only.'),
('ML warning','Exclude adjusted_risk_score, risk_level, and action_status from model features to avoid target leakage. Consider also excluding inherent/residual scores when testing whether ML adds value beyond the scoring baseline.'),
('Reproducibility','Random seed: 42. Generator script is supplied alongside the datasets.')]
for i,(k,v) in enumerate(notes,3): ws.cell(i,1,k).font=Font(bold=True,color='666666'); ws.cell(i,2,v); ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=6)
ws.column_dimensions['A'].width=22
for c in range(2,7): ws.column_dimensions[get_column_letter(c)].width=22

# dictionary
wd=wb.create_sheet('Data Dictionary'); wd.sheet_view.showGridLines=False
wd.append(['dataset','field','type','role','description'])
desc={
'risk_id':'Unique synthetic risk identifier','risk_level':'Synthetic classification target: Low, Medium, High, or Critical','adjusted_risk_score':'Latent synthetic expert score used to create the target; exclude from ML features','residual_risk_score':'Inherent risk reduced by modeled control effectiveness','control_effectiveness':'Modeled effectiveness from 0 to 1','synthetic_anomaly_label':'Injected anomaly indicator used only to evaluate anomaly detection','linked_risk_id':'Foreign key to the risk register','failed_samples':'Number of control-test samples that failed'}
for ds,hs in [('risk_register.csv',risk_headers),('control_register.csv',control_headers),('access_activity.csv',access_headers)]:
    for h in hs:
        role='Target' if h in ['risk_level','synthetic_anomaly_label'] else ('Identifier' if h.endswith('_id') else ('Derived' if h in ['inherent_risk_score','residual_risk_score','adjusted_risk_score','action_status','test_status'] else 'Feature'))
        typ='numeric' if h in ['likelihood','impact','asset_criticality','data_sensitivity','vulnerability_count','critical_vulnerability_count','past_incident_count','policy_violation_count','audit_finding_count','third_party_dependency','number_of_controls','control_effectiveness','control_test_pass_rate','days_since_control_test','open_control_deficiencies','treatment_due_days','risk_accepted','days_open','inherent_risk_score','residual_risk_score','adjusted_risk_score','automated_control','sample_size','failed_samples','privileged_access','successful_logins','failed_logins','after_hours_events','sensitive_records_accessed','distinct_locations','synthetic_anomaly_label'] else ('date/datetime' if 'date' in h or 'timestamp' in h else 'categorical/text')
        wd.append([ds,h,typ,role,desc.get(h,h.replace('_',' ').capitalize())])
for c in wd[1]: c.fill=PatternFill('solid',fgColor=navy); c.font=Font(color='FFFFFF',bold=True)
wd.freeze_panes='A2'; wd.auto_filter.ref=f'A1:E{wd.max_row}'
for i,w in enumerate([24,32,18,16,72],1): wd.column_dimensions[get_column_letter(i)].width=w

path=OUT/'grc_project_datasets.xlsx'; wb.save(path)
print('created',path)
print('risk levels',Counter(r[-2] for r in risk_rows))
print('anomalies',sum(r[-1] for r in access_rows))
